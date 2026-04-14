"""
app.py  —  KTS Excel Export Backend (single-file version)
==========================================================
Everything merged into one file — no subfolders needed.
Just upload this app.py to GitHub and Render will run it.
"""

from __future__ import annotations

import io
import logging
import os
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Optional

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.fills import FILL_SOLID

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
)
logger = logging.getLogger(__name__)

# ── App setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")
CORS(app, resources={r"/api/*": {"origins": ALLOWED_ORIGIN}})

BASE_DIR       = Path(__file__).parent
TEMPLATES_DIR  = BASE_DIR / "templates_excel"

# Search for templates in multiple locations — handles different upload arrangements
def _find_template(short_name: str, long_patterns: list) -> Path:
    """Find a template file by trying several possible locations/names."""
    # 1. Standard location
    p = TEMPLATES_DIR / short_name
    if p.is_file(): return p
    # 2. Root folder with short name
    p = BASE_DIR / short_name
    if p.is_file(): return p
    # 3. Root folder with long/original name
    for pattern in long_patterns:
        for f in BASE_DIR.glob(pattern):
            return f
    return TEMPLATES_DIR / short_name  # fallback (will raise FileNotFoundError)

HA0935_TEMPLATE = _find_template(
    "HA0935_template.xlsx",
    ["HA0935*.xlsx", "HA0935_*.xlsx"]
)
HV0713_TEMPLATE = _find_template(
    "HV0713_template.xlsx",
    ["HV0713*.xlsx", "HV0713_*.xlsx"]
)

# ── Yellow-cell helpers ───────────────────────────────────────────────────────
_YELLOW_SUFFIX = "FFFF99"

def is_yellow(cell) -> bool:
    try:
        return (
            cell.fill is not None
            and cell.fill.fill_type == FILL_SOLID
            and cell.fill.fgColor.rgb.upper().endswith(_YELLOW_SUFFIX)
        )
    except Exception:
        return False

def _letter(col: int) -> str:
    s = ""
    while col > 0:
        col, r = divmod(col - 1, 26)
        s = chr(65 + r) + s
    return s

def safe_write(ws, row: int, col: int, value: Any) -> bool:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return False
    if isinstance(cell.value, str) and cell.value.startswith("="):
        raise RuntimeError(f"Formula-overwrite blocked at {ws.title}!{_letter(col)}{row}")
    if not is_yellow(cell):
        logger.error("SAFETY: %s!%s%d is not yellow — write rejected.", ws.title, _letter(col), row)
        return False
    cell.value = value
    return True

# ── Template loader ───────────────────────────────────────────────────────────
def load_template(template_path: Path):
    if not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    tmp_path = Path(tmp.name)
    shutil.copy2(template_path, tmp_path)
    wb = openpyxl.load_workbook(str(tmp_path), data_only=False, keep_vba=False)
    return wb, tmp_path

def finish(wb, tmp_path: Path) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    try:
        tmp_path.unlink()
    except Exception:
        pass
    return data

# ── HA0935 writer ─────────────────────────────────────────────────────────────
HA_ROUTE_NAMES    = ["Khan1","Khan2","Khan3","Khan4","Khan5","Khan6","Khan7","Khan8","Khan9","Khan10"]
HA_OVO_FIRST_ROW  = 10
HA_SSO_FIRST_ROW  = 41
HA_DAY_COL_C      = 3

def build_ha0935(template_path: Path, data: dict) -> bytes:
    wb, tmp_path = load_template(template_path)

    # Route Detail
    ws = wb["Route Detail"]
    ws["A2"].value = data["month_date"]

    ovo = data.get("ovo", {})
    for ri, route_name in enumerate(HA_ROUTE_NAMES):
        am_row = HA_OVO_FIRST_ROW + ri * 2
        pm_row = am_row + 1
        route_ovo = ovo.get(route_name, {})
        for period, row_num in [("AM", am_row), ("PM", pm_row)]:
            for day, value in route_ovo.get(period, {}).items():
                if value is None:
                    continue
                col = HA_DAY_COL_C + (int(day) - 1)
                if col > 33:
                    continue
                safe_write(ws, row_num, col, value)

    sso = data.get("sso", {})
    for ri, route_name in enumerate(HA_ROUTE_NAMES):
        row_num = HA_SSO_FIRST_ROW + ri
        for day, miles in sso.get(route_name, {}).items():
            col = HA_DAY_COL_C + (int(day) - 1)
            if col > 33:
                continue
            safe_write(ws, row_num, col, float(miles))

    if data.get("vehicle_unit_rate") is not None:
        ws["C35"].value = float(data["vehicle_unit_rate"])
    if data.get("contract_mile_rate") is not None:
        ws["C56"].value = float(data["contract_mile_rate"])

    # Daily Attendance Report
    ws2 = wb["Daily Attendance Report"]
    ws2["A2"].value = data["month_date"]
    for row_str, day_codes in data.get("attendance", {}).items():
        row_num = int(row_str)
        for day_str, code in day_codes.items():
            col = 7 + (int(day_str) - 1)   # col G = 7 = day 1
            if col > 37:
                continue
            val = int(code) if str(code).lstrip("-").isdigit() else str(code).upper()
            safe_write(ws2, row_num, col, val)

    return finish(wb, tmp_path)

# ── HV0713 writer ─────────────────────────────────────────────────────────────
HV_OVO_ROW_MAP = {
    "6770": {"AM": 9,  "PM": 10},
    "6771": {"AM": 11, "PM": 12},
}
HV_SSO_ROW_MAP = {"6770": 24, "6771": 25}
HV_DAY_COL_C   = 3

def build_hv0713(template_path: Path, data: dict) -> bytes:
    wb, tmp_path = load_template(template_path)

    # Route Detail
    ws = wb["Route Detail"]
    ws["A2"].value = data["month_date"]

    ovo = data.get("ovo", {})
    for route_name, periods in ovo.items():
        rows = HV_OVO_ROW_MAP.get(str(route_name), {})
        for period, day_values in periods.items():
            row_num = rows.get(period.upper())
            if row_num is None:
                continue
            for day, value in day_values.items():
                if value is None:
                    continue
                col = HV_DAY_COL_C + (int(day) - 1)
                if col > 33:
                    continue
                safe_write(ws, row_num, col, value)

    sso = data.get("sso", {})
    for route_name, daily_miles in sso.items():
        row_num = HV_SSO_ROW_MAP.get(str(route_name))
        if row_num is None:
            continue
        for day, miles in daily_miles.items():
            col = HV_DAY_COL_C + (int(day) - 1)
            if col > 33:
                continue
            safe_write(ws, row_num, col, float(miles))

    if data.get("vehicle_unit_rate") is not None:
        ws["C18"].value = float(data["vehicle_unit_rate"])
    if data.get("contract_mile_rate") is not None:
        ws["C31"].value = float(data["contract_mile_rate"])

    # Daily Attendance Report
    ws2 = wb["Daily Attendance Report"]
    ws2["A2"].value = data["month_date"]
    for row_str, day_codes in data.get("attendance", {}).items():
        row_num = int(row_str)
        for day_str, code in day_codes.items():
            col = 7 + (int(day_str) - 1)
            if col > 37:
                continue
            val = int(code) if str(code).lstrip("-").isdigit() else str(code).upper()
            safe_write(ws2, row_num, col, val)

    return finish(wb, tmp_path)

# ── Request helpers ───────────────────────────────────────────────────────────
def _parse_date(raw: str) -> date:
    return datetime.strptime(raw.strip(), "%Y-%m-%d").date()

def _parse_body(body: dict) -> dict:
    """Normalise JSON payload — convert string day keys to int, etc."""
    month_date = _parse_date(body.get("month_date", ""))
    if month_date.day != 1:
        raise ValueError(f"month_date must be the 1st of the month, got {month_date}")

    def norm_ovo(raw):
        result = {}
        for route, periods in (raw or {}).items():
            result[str(route)] = {
                p.upper(): {int(k): int(v) for k, v in days.items()}
                for p, days in periods.items()
            }
        return result

    def norm_sso(raw):
        return {
            str(route): {int(k): float(v) for k, v in days.items()}
            for route, days in (raw or {}).items()
        }

    def norm_att(raw):
        result = {}
        for row_str, days in (raw or {}).items():
            result[row_str] = {
                str(d): (int(c) if str(c).lstrip("-").isdigit() else str(c))
                for d, c in days.items()
            }
        return result

    return {
        "month_date":          month_date,
        "ovo":                 norm_ovo(body.get("ovo", {})),
        "sso":                 norm_sso(body.get("sso", {})),
        "attendance":          norm_att(body.get("attendance", {})),
        "vehicle_unit_rate":   body.get("vehicle_unit_rate"),
        "contract_mile_rate":  body.get("contract_mile_rate"),
    }

def _xlsx_response(xlsx_bytes: bytes, filename: str):
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok",
        "ha0935_template": HA0935_TEMPLATE.is_file(),
        "hv0713_template": HV0713_TEMPLATE.is_file(),
    })

@app.route("/api/export/ha0935", methods=["POST"])
def export_ha0935():
    if not request.is_json:
        return jsonify({"error": "Content-Type must be application/json"}), 400
    try:
        data = _parse_body(request.get_json(silent=True) or {})
    except Exception as e:
        return jsonify({"error": str(e)}), 422
    try:
        xlsx_bytes = build_ha0935(HA0935_TEMPLATE, data)
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("HA0935 export error")
        return jsonify({"error": str(e)}), 500
    filename = f"HA0935_Khan_{data['month_date'].strftime('%b%Y')}.xlsx"
    logger.info("HA0935 exported: %s", filename)
    return _xlsx_response(xlsx_bytes, filename)

@app.route("/api/export/hv0713", methods=["POST"])
def export_hv0713():
    if not request.is_json:
        return jsonify({"error": "Content-Type must be application/json"}), 400
    try:
        data = _parse_body(request.get_json(silent=True) or {})
    except Exception as e:
        return jsonify({"error": str(e)}), 422
    try:
        xlsx_bytes = build_hv0713(HV0713_TEMPLATE, data)
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("HV0713 export error")
        return jsonify({"error": str(e)}), 500
    filename = f"HV0713_Priority_{data['month_date'].strftime('%b%Y')}.xlsx"
    logger.info("HV0713 exported: %s", filename)
    return _xlsx_response(xlsx_bytes, filename)

# ── Dev server ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
